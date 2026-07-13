"""Synchronize the account workbook for the July 2026 competitor update."""

from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = PROJECT_ROOT / "files" / "social_media_accounts.xlsx"

NEW_ACCOUNT_ROWS = [
    [
        "Kimi Work",
        "@Kimi_Moonshot",
        "—",
        "—",
        "@kimiai.official",
        "—",
        "—",
        "—",
        "—",
        "—",
        "Kimi智能助手 UID 3546607314274766",
        "—",
    ],
    ["TRAE Work", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—"],
    [
        "MiniMax Code",
        "@MiniMax__AI",
        "—",
        "—",
        "—",
        "—",
        "—",
        "—",
        "—",
        "—",
        "MiniMax_稀宇极智 UID 3546878859807653",
        "—",
    ],
    [
        "Marvis",
        "—",
        "—",
        "—",
        "—",
        "—",
        "—",
        "—",
        "—",
        "—",
        "Marvis马维斯 UID 3706967468018619",
        "—",
    ],
    [
        "QClaw",
        "—",
        "—",
        "—",
        "—",
        "—",
        "—",
        "—",
        "—",
        "—",
        "腾讯电脑管家-管哥 团队账号 UID 626140785",
        "—",
    ],
    ["阶跃 AI", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—"],
]


def delete_product_rows(
    worksheet, products: set[str], product_column: int = 1
) -> None:
    for row in range(worksheet.max_row, 1, -1):
        product = str(worksheet.cell(row, product_column).value or "").strip()
        if product in products:
            worksheet.delete_rows(row)


def append_styled_row(worksheet, values: list, template_row: int) -> None:
    target_row = worksheet.max_row + 1
    for column, value in enumerate(values, 1):
        source = worksheet.cell(template_row, column)
        target = worksheet.cell(target_row, column, value)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[
        template_row
    ].height


def remove_incomplete_people_rows(worksheet) -> None:
    """Remove stale hyperlink-only rows left by Excel row deletion."""
    for row in range(worksheet.max_row, 1, -1):
        person = str(worksheet.cell(row, 1).value or "").strip()
        product = str(worksheet.cell(row, 2).value or "").strip()
        if person and product:
            continue
        for cell in worksheet[row]:
            cell.hyperlink = None
        worksheet.delete_rows(row)


def main() -> None:
    workbook = load_workbook(WORKBOOK_PATH)
    accounts = workbook["竞品官方账号总览"]
    people = workbook["关键人物账号"]
    coverage = workbook["平台覆盖统计"]

    changed_products = {"n8n", *(row[0] for row in NEW_ACCOUNT_ROWS)}
    delete_product_rows(accounts, changed_products)
    delete_product_rows(coverage, changed_products)
    delete_product_rows(people, {"n8n"}, product_column=2)
    remove_incomplete_people_rows(people)

    account_template = accounts.max_row
    coverage_template = coverage.max_row
    for account_row in NEW_ACCOUNT_ROWS:
        append_styled_row(accounts, account_row, account_template)
        marks = [
            "✓" if value not in (None, "", "—") else "✗"
            for value in account_row[1:]
        ]
        coverage_row = [
            account_row[0],
            *marks,
            sum(mark == "✓" for mark in marks),
        ]
        append_styled_row(coverage, coverage_row, coverage_template)

    for worksheet in (accounts, people, coverage):
        worksheet.auto_filter.ref = worksheet.dimensions

    workbook.save(WORKBOOK_PATH)
    print(
        f"updated {WORKBOOK_PATH}: accounts={accounts.max_row - 1}, "
        f"people={people.max_row - 1}, coverage={coverage.max_row - 1}"
    )


if __name__ == "__main__":
    main()
